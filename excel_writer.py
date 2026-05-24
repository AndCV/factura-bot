"""
excel_writer.py — Crea o actualiza el Excel mensual de facturas.
"""
import os
from datetime import date

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

import config

HEADERS    = ["N° Factura", "Fecha", "Emisor", "Descripción", "Tipo", "Otros", "Monto", "IVA", "Fuente"]
COL_WIDTHS = [28, 14, 30, 36, 16, 20, 14, 12, 8]

COLOR_HEADER_BG = "1F4E79"
COLOR_HEADER_FG = "FFFFFF"
COLOR_ROW_ALT   = "D6E4F0"


def _ruta_excel():
    mes = date.today().strftime("%Y-%m")
    nombre = f"{config.EXCEL_BASE_NAME}_{mes}.xlsx"
    os.makedirs(config.EXCEL_FOLDER, exist_ok=True)
    return os.path.join(config.EXCEL_FOLDER, nombre)


def _thin_border():
    thin = Side(style="thin", color="AAAAAA")
    return Border(left=thin, right=thin, top=thin, bottom=thin)


def _crear_hoja_nueva(wb):
    ws = wb.active
    ws.title = "Facturas"

    for col, (header, width) in enumerate(zip(HEADERS, COL_WIDTHS), start=1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font      = Font(bold=True, color=COLOR_HEADER_FG, name="Arial", size=11)
        cell.fill      = PatternFill("solid", fgColor=COLOR_HEADER_BG)
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border    = _thin_border()
        ws.column_dimensions[get_column_letter(col)].width = width

    ws.row_dimensions[1].height = 22
    ws.freeze_panes = "A2"

    ws2 = wb.create_sheet("Resumen")
    ws2["A1"] = "Resumen mensual"
    ws2["A1"].font = Font(bold=True, size=13, name="Arial")
    ws2["A3"] = "Total facturas"
    ws2["A4"] = "Total monto"
    ws2["A5"] = "Total IVA"
    for cell in [ws2["A3"], ws2["A4"], ws2["A5"]]:
        cell.font = Font(name="Arial", bold=True)
    ws2["B3"] = "=COUNTA(Facturas!A2:A10000)"
    ws2["B4"] = "=SUM(Facturas!G2:G10000)"
    ws2["B5"] = "=SUM(Facturas!H2:H10000)"
    ws2["B4"].number_format = '₡#,##0.00'
    ws2["B5"].number_format = '₡#,##0.00'
    ws2.column_dimensions["A"].width = 20
    ws2.column_dimensions["B"].width = 18

    return ws


def _siguiente_fila(ws):
    row = 2
    while ws.cell(row=row, column=1).value is not None:
        row += 1
    return row


def _formatear_fila(ws, row):
    fill = PatternFill("solid", fgColor=COLOR_ROW_ALT if row % 2 == 0 else "FFFFFF")
    for col in range(1, len(HEADERS) + 1):
        cell = ws.cell(row=row, column=col)
        cell.fill      = fill
        cell.border    = _thin_border()
        cell.font      = Font(name="Arial", size=10)
        cell.alignment = Alignment(vertical="center")
        if col == 7:   # Monto
            cell.number_format = '₡#,##0.00'
            cell.alignment = Alignment(horizontal="right")
        if col == 8:   # IVA
            cell.number_format = '₡#,##0.00'
            cell.alignment = Alignment(horizontal="right")


def guardar_facturas(facturas: list[dict]):
    ruta = _ruta_excel()
    es_nuevo = not os.path.exists(ruta)

    if es_nuevo:
        wb = openpyxl.Workbook()
        ws = _crear_hoja_nueva(wb)
    else:
        wb = openpyxl.load_workbook(ruta)
        ws = wb["Facturas"]

    agregadas = 0
    for f in facturas:
        if not f:
            continue
        row = _siguiente_fila(ws)
        ws.cell(row=row, column=1, value=f.get("numero", ""))
        ws.cell(row=row, column=2, value=f.get("fecha", ""))
        ws.cell(row=row, column=3, value=f.get("emisor", ""))
        ws.cell(row=row, column=4, value=f.get("descripcion", ""))
        ws.cell(row=row, column=5, value=f.get("tipo", ""))
        ws.cell(row=row, column=6, value=f.get("otros", ""))
        ws.cell(row=row, column=7, value=f.get("monto", 0))
        ws.cell(row=row, column=8, value=f.get("iva", 0))
        ws.cell(row=row, column=9, value=f.get("fuente", ""))
        _formatear_fila(ws, row)
        agregadas += 1

    wb.save(ruta)
    print(f"\n✅ {agregadas} factura(s) guardadas en: {ruta}")
    return ruta
